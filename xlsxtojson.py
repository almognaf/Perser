import openpyxl
import json
import time

""" get_report()
function that receiving the number of the current report's row and the sheet
and returns only the parameters we need from the "report" area. """


def get_report(current_row, sheets):
    tube_num = None
    thermal = None
    deorb = None
    found = False
    i = 1
    # while not reached to next sample
    while not found and not str(sheets['A' + str(int(current_row) + i)].value).__contains__(
            "Sample"):
        #  coord is parameter that holding the
        coord = 'C' + str(int(current_row) + i)
        if str(sheets[coord].value).__contains__("Tube Number"):
            tube_num = int(sheets['I' + str(int(current_row) + i)].value)
        if str(sheets[coord].value).__contains__("Thermal Cycles"):
            thermal = int(sheets['I' + str(int(current_row) + i)].value)
        if str(sheets[coord].value).__contains__("Desorb Start Time"):
            deorb = str(sheets['I' + str(int(current_row) + i)].value)

        i += 1
        if int(current_row) + i == sheets.max_row:
            break

    return tube_num, thermal, deorb


def create_json(to_json):
    with open('C:/Users/alnaf/Perser/Jsons/' + 'result_' + str(time.time()) + '.json', 'w') as fp:
        json.dump(to_json, fp, indent=4)
        fp.close()


def xlsx_to_json(xlsx_path):
    st = time.time()
    path = xlsx_path
    # path = 'Reports/rawdata.xlsx'

    # load excel file
    book = openpyxl.load_workbook(path)

    # select the sheet
    sheet = book['SeqReport']

    sample_col = 'A'
    sheet_curr_row = '1'
    # sample_row_count to handle last sample case and for make sure we in range
    samples = []
    curr_sample = {}
    sample_header_row = 1

    # while loop to ensure we start the program at the first sample
    while not str(sheet[sample_col + str(sample_header_row)].value).__contains__("Sample"):
        sample_header_row += 1

    #  while loop to go over the samples in the Excel sheet without overflowing
    while str(sheet[sample_col + str(sample_header_row)].value).__contains__("Sample") and int(
            sheet_curr_row) < sheet.max_row:
        sheet_curr_row = str(sample_header_row + 1)
        while not str(sheet[sample_col + str(sheet_curr_row)].value).__contains__("Sample") and int(
                sheet_curr_row) < sheet.max_row:

            #  if statement for parsing relevant data inside "File" area if we there
            if str(sheet['B' + sheet_curr_row].value).__contains__("File"):
                #  taking the value inside col 'F' at our curr_row
                file = sheet['F' + sheet_curr_row].value
                curr_sample["File"] = file

            #  if statement for parsing relevant data inside "Method" area if we there
            if str(sheet['B' + sheet_curr_row].value).__contains__("Method:"):
                #  taking the value inside col 'F' at our curr_row
                method = sheet['F' + sheet_curr_row].value
                curr_sample["Method"] = method

            #  if statement for parsing relevant data inside "Markes" area if we there
            if str(sheet['B' + sheet_curr_row].value).__contains__("Markes"):
                # tube start after 2 rows
                fixed_row = str(int(sheet_curr_row) + 2)
                tube = int(sheet['I' + fixed_row].value)
                curr_sample["Tube"] = tube

            #  if statement for parsing relevant data inside "Report" area if we there
            if str(sheet['B' + sheet_curr_row].value).__contains__("Report"):
                #  using the get_report function we can fetch the data we need from "Report" section
                tube_number, thermal_cycles, desorb_start_time = get_report(sheet_curr_row, sheet)
                curr_sample["Tube_Number"] = tube_number
                curr_sample["Thermal_Cycles"] = thermal_cycles
                curr_sample["Desorb_Start_Time"] = desorb_start_time
            sheet_curr_row = str(int(sheet_curr_row) + 1)

        #  Append a copy of the last dict to Samples[] for not loosing it in the next loop
        samples.append(curr_sample.copy())

        """
                   iterating the general sheet row (sheetCurrRow) in 1 every time we not in the inside
                   Parsing while loop until one of the following conditions will happen (pseudo code):
                   sheetCurrRow.value.contains("Sample") = parse new sample data
                   or
                   sheetCurrRow !< sheet.max_row = end of excel sheet and program
        """

        sample_header_row = int(sheet_curr_row)

    openpyxl.Workbook.close(book)
    #  taking the dict array made from the Excel sheet we received and creating the Json we need
    create_json(samples)
    et = time.time()

    # get the execution time
    elapsed_time = et - st
    print('Execution time:', elapsed_time, 'seconds')